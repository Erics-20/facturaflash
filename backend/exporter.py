"""Exporta cabecera + productos a Excel (.xlsx) y CSV."""

from __future__ import annotations
import io
import pandas as pd

_CABECERA_LABELS: list[tuple[str, str]] = [
    ("proveedor",    "Proveedor"),
    ("ruc",          "RUC"),
    ("serie_numero", "Serie-Número"),
    ("fecha",        "Fecha"),
    ("total",        "Total (S/)"),
]

_COLUMN_LABELS = {
    "descripcion":     "Descripción",
    "cantidad":        "Cantidad",
    "precio_unitario": "Precio Unit. (S/)",
    "subtotal":        "Subtotal (S/)",
}


def to_excel_bytes(
    df: pd.DataFrame,
    cabecera: dict | None = None,
    sheet_name: str = "Factura",
) -> bytes:
    """
    Genera un .xlsx con cabecera arriba y productos debajo.

    Layout:
        [Etiqueta bold] [Valor]   ← una fila por campo de cabecera
        (fila en blanco)
        [Encabezados de tabla bold]
        [filas de productos]
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAD3")   # verde suave

    current_row = 1

    # ── Bloque de cabecera ──────────────────────────────────────────────────
    if cabecera:
        for key, label in _CABECERA_LABELS:
            value = cabecera.get(key, "")
            if value is None:
                value = ""
            label_cell = ws.cell(row=current_row, column=1, value=f"{label}:")
            label_cell.font = bold
            label_cell.alignment = Alignment(horizontal="right")
            value_cell = ws.cell(row=current_row, column=2, value=value)
            if key == "total":
                value_cell.number_format = "#,##0.00"
            current_row += 1

        # Fila en blanco de separación
        current_row += 1

    # ── Encabezados de productos ────────────────────────────────────────────
    col_headers = [_COLUMN_LABELS.get(c, c) for c in df.columns]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
    current_row += 1

    # ── Filas de productos ──────────────────────────────────────────────────
    numeric_headers = {"Cantidad", "Precio Unit. (S/)", "Subtotal (S/)"}
    for _, row_data in df.iterrows():
        for col_idx, (col_name, value) in enumerate(
            zip(df.columns, row_data), start=1
        ):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if _COLUMN_LABELS.get(col_name) in numeric_headers:
                cell.number_format = "#,##0.00"
        current_row += 1

    # ── Ajustar anchos de columna ───────────────────────────────────────────
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 4, 55
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_csv_string(
    df: pd.DataFrame,
    cabecera: dict | None = None,
) -> str:
    """
    Genera un CSV con cabecera arriba y productos debajo.

    Retorna un string; encódalo a bytes con .encode('utf-8-sig') al descargarlo.
    """
    lines: list[str] = []

    if cabecera:
        for key, label in _CABECERA_LABELS:
            value = cabecera.get(key, "")
            if value is None:
                value = ""
            # Escapar comas dentro del valor
            value_str = str(value).replace('"', '""')
            if "," in value_str:
                value_str = f'"{value_str}"'
            lines.append(f"{label}:,{value_str}")
        lines.append("")   # fila en blanco

    # Cabecera y datos de productos
    export_df = df.rename(columns=_COLUMN_LABELS)
    lines.append(export_df.to_csv(index=False).rstrip("\n"))

    return "\n".join(lines)
